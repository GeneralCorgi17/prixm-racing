#!/usr/bin/env python3
"""
Prixm Qualifying Picks Excel Exporter.

Generates qualifying_picks.xlsx — all dates from START_DATE onwards.
Results (W/L/V) and Price (SP) are auto-filled from results_history.json.
Future races stay blank until results_fetcher.py runs and triggers a rebuild.

Filters:
  Before 2026-05-29: gap >= 17, non-handicap, England only, price > 2.00
  From  2026-05-29:  gap >= 18, score >= 74, non-handicap, England only, price > 2.00

Usage:
    python qualifying_exporter.py          # rebuild combined file
    python qualifying_exporter.py --scan   # same (explicit)
"""

import argparse
import datetime
import glob
import json
import os
import re
import sys

OUTPUT_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'output', 'qualifying_picks.xlsx')
START_DATE        = '2026-04-08'
MIN_GAP           = 17    # legacy threshold (before NEW_THRESHOLD_DATE)
MIN_GAP_NEW       = 18    # from 2026-05-29 onwards
MIN_SCORE_NEW     = 74    # from 2026-05-29 onwards
NEW_THRESHOLD_DATE = '2026-05-29'

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print('ERROR: openpyxl not installed.  Run: pip install openpyxl')
    sys.exit(1)

# ── Colours ───────────────────────────────────────────────────────────────────
C_TITLE_BG = '12121f'
C_TITLE_FG = 'c9a84c'
C_HDR_BG   = '1a1a2e'
C_HDR_FG   = 'c9a84c'
C_SUB_FG   = '888888'
C_ALT      = 'f5f5f0'
C_PRICE_BG = 'fffacd'
C_WIN_BG   = 'd4edda'
C_LOSS_BG  = 'f8d7da'
C_VOID_BG  = 'fff3cd'
C_BORDER       = 'd0d0d0'
C_FOOTER       = '999999'
C_LABEL_GOLDEN = 'c9a84c'   # gold text
C_LABEL_SILVER = 'a8a9ad'   # silver text
C_LABEL_BRONZE = 'b87333'   # bronze/copper text

# ── Columns ───────────────────────────────────────────────────────────────────
COLUMNS = [
    ('Date',       12),
    ('Venue',      18),
    ('Time',        8),
    ('Horse',      24),
    ('Score',       8),
    ('Gap',         8),
    ('Sub Label',  14),
    ('Jockey',     22),
    ('Trainer',    24),
    ('Price',      10),
    ('Decimal',     9),
    ('Result',     10),
    ('Net (£)',    10),
]
N_COLS   = len(COLUMNS)
LAST_COL = get_column_letter(N_COLS)

COL_DATE      = 1
COL_VENUE     = 2
COL_TIME      = 3
COL_HORSE     = 4
COL_SCORE     = 5
COL_GAP       = 6
COL_SUB_LABEL = 7
COL_JOCKEY    = 8
COL_TRAINER   = 9
COL_PRICE     = 10
COL_DECIMAL   = 11
COL_RESULT    = 12
COL_NET       = 13

STAKE_CELL  = '$O$2'   # editable stake input — user changes this

EXCLUDED_VENUES = [
    # Asia / HK / Middle East
    'happy valley', 'sha tin', 'meydan', 'abu dhabi', 'bahrain',
    'hanshin', 'chukyo', 'fukushima', 'nakayama', 'tokyo', 'niigata', 'kyoto',
    # Australasia
    'rosehill', 'flemington', 'morphettville', 'eagle farm', 'doomben',
    'hawkesbury', 'ascot aus', 'gold coast', 'northam', 'scone',
    # South Africa
    'turffontein', 'scottsville', 'greyville',
    # USA / Canada
    'oaklawn park', 'gulfstream park', 'keeneland', 'aqueduct', 'santa anita',
    'belmont park', 'belmont at the big a', 'churchill downs', 'laurel park',
    'lone star park', 'percy warner', 'middleburg', 'great meadow', 'woodbine', 'saratoga',
    'hipodromo',
    # South America
    'cidade jardim', 'monterrico', 'gavea', 'palermo',
    # France
    'saint cloud', 'longchamp', 'deauville', 'chantilly', 'auteuil',
    'compiegne', 'toulouse', 'bordeaux', 'les landes', 'nantes', 'vichy', 'le lion',
    # Germany / Italy
    'san siro', 'munich', 'dusseldorf', 'krefeld', 'hoppegarten', 'dortmund', 'baden baden',
    'cologne', 'koln', 'firenze', 'randwick',
    # Ireland (all)
    'punchestown', 'leopardstown', 'curragh', 'naas', 'cork',
    'killarney', 'gowran park', 'roscommon', 'navan', 'limerick',
    'clonmel', 'wexford', 'tramore', 'kilbeggan', 'ballinrobe', 'sligo',
    'down royal', 'bellewstown', 'downpatrick', 'dundalk', 'tipperary',
    'fairyhouse', 'galway', 'laytown',
    # Meta
    'free to air', 'scoop', 'worldwide stakes', 'world pool',
]


# ── Name / venue normalisation ────────────────────────────────────────────────

def _norm_venue(name):
    """Strip whitespace, country suffix, lowercase for matching."""
    v = re.sub(r'\s+', ' ', str(name)).strip().lower()
    v = re.sub(r'\s*\([a-z]{2,3}\)\s*$', '', v).strip()   # remove (IRE), (FR) etc
    v = v.replace('-', ' ')
    return v


def _norm_name(name):
    return re.sub(r'[^a-z]', '', str(name).lower())


_IRE_TRACKS = {
    'curragh', 'leopardstown', 'fairyhouse', 'punchestown', 'galway',
    'naas', 'cork', 'tipperary', 'gowran park', 'killarney', 'listowel',
    'dundalk', 'navan', 'tramore', 'sligo', 'roscommon', 'ballinrobe',
    'clonmel', 'limerick', 'thurles', 'wexford', 'bellewstown',
    'down royal', 'downpatrick', 'laytown',
}

def _is_ireland(venue_name):
    v = venue_name.lower()
    if '(ire)' in v:
        return True
    return _norm_venue(venue_name) in _IRE_TRACKS


def _is_excluded(venue_name):
    if _is_ireland(venue_name):
        return True
    v = _norm_venue(venue_name)
    return any(ex in v for ex in EXCLUDED_VENUES)


# ── Results history lookup ────────────────────────────────────────────────────

def _build_history_lookup(output_dir):
    """
    Returns {(date, norm_venue, time): {norm_horse: runner_dict}}.
    norm_horse key uses _norm_name for fuzzy matching.
    """
    path = os.path.join(output_dir, 'results_history.json')
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            history = json.load(f)
    except Exception:
        return {}

    lookup = {}
    for race in history.get('races', []):
        key = (race.get('date', ''), _norm_venue(race.get('venue', '')), race.get('time', ''))
        lookup[key] = {
            _norm_name(r.get('name', '')): r
            for r in race.get('results', [])
        }
    return lookup


def _resolve_result(qualifier, history_lookup):
    """
    Look up the qualifying horse in results_history.
    Returns (result_str, price_str) or ('', '') if not found yet.
      result_str: 'W', 'L', or 'V'
      price_str:  fractional SP string e.g. '11/4', or decimal str, or ''
    """
    race_key = (qualifier['date'], _norm_venue(qualifier['venue']), qualifier['time'])
    race_runners = history_lookup.get(race_key)
    if not race_runners:
        return '', ''

    norm_h = _norm_name(qualifier['horse'])
    runner = race_runners.get(norm_h)

    # Fuzzy fallback: substring match
    if not runner:
        for k, v in race_runners.items():
            if norm_h in k or k in norm_h:
                runner = v
                break

    if not runner:
        return '', ''

    if runner.get('non_runner') or runner.get('pull_out'):
        return 'V', ''

    fp = runner.get('finish_pos')
    if fp == 1:
        result = 'W'
    elif fp:
        result = 'L'
    else:
        return 'V', ''   # in race but no finish pos (PU/fell/DSQ/etc.)

    # Price: prefer fractional sp_raw, fall back to decimal sp
    sp_raw = runner.get('sp_raw') or ''
    sp_dec = runner.get('sp')
    if sp_raw:
        price = sp_raw
    elif sp_dec is not None:
        price = str(sp_dec)
    else:
        price = ''

    return result, price


# ── Core filter ───────────────────────────────────────────────────────────────

BRONZE_GOING = {'good', 'good to firm'}
AW_GOING     = {'standard', 'standard to slow'}
NH_TYPES     = {'Chase', 'Hurdle', 'NH Flat', 'Bumper'}


def collect_qualifiers(data, date_str):
    qualifiers = []
    use_new = date_str >= NEW_THRESHOLD_DATE
    min_gap = MIN_GAP_NEW if use_new else MIN_GAP

    for venue_name, venue in data.get('venues', {}).items():
        if _is_excluded(venue_name):
            continue
        for race in venue.get('races', []):
            if race.get('handicap', False):
                continue
            runners = sorted(
                race.get('runners', []),
                key=lambda r: r.get('score', 0),
                reverse=True
            )
            if len(runners) < 2:
                continue
            gap       = round(runners[0].get('score', 0) - runners[1].get('score', 0), 1)
            top       = runners[0]
            score     = top.get('score', 0)
            race_type = race.get('race_type', '')
            going     = (race.get('going') or '').lower().strip()

            # ── Golden ────────────────────────────────────────────────
            if gap >= min_gap and (not use_new or score >= MIN_SCORE_NEW):
                sub_label = 'Golden_NH' if race_type in NH_TYPES else 'Golden_Flat'
                qualifiers.append({
                    'date':      date_str,
                    'venue':     venue_name,
                    'time':      race.get('time', ''),
                    'horse':     top.get('name', ''),
                    'score':     score,
                    'gap':       gap,
                    'sub_label': sub_label,
                    'pick_type': 'golden',
                    'jockey':    top.get('jockey', ''),
                    'trainer':   top.get('trainer', ''),
                })

            # ── Silver (log only — paper track until 30+ results) ────
            elif 10 <= gap < 12 and score >= 72 and going not in AW_GOING:
                sub_label = 'Silver_NH' if race_type in NH_TYPES else 'Silver_Flat'
                qualifiers.append({
                    'date':      date_str,
                    'venue':     venue_name,
                    'time':      race.get('time', ''),
                    'horse':     top.get('name', ''),
                    'score':     score,
                    'gap':       gap,
                    'sub_label': sub_label,
                    'pick_type': 'silver',
                    'jockey':    top.get('jockey', ''),
                    'trainer':   top.get('trainer', ''),
                })

            # ── Bronze (log only — paper track until 30+ results) ─────
            elif 8 <= gap < 10 and score >= 74 and going in BRONZE_GOING:
                sub_label = 'Bronze_NH' if race_type in NH_TYPES else 'Bronze_Flat'
                qualifiers.append({
                    'date':      date_str,
                    'venue':     venue_name,
                    'time':      race.get('time', ''),
                    'horse':     top.get('name', ''),
                    'score':     score,
                    'gap':       gap,
                    'sub_label': sub_label,
                    'pick_type': 'bronze',
                    'jockey':    top.get('jockey', ''),
                    'trainer':   top.get('trainer', ''),
                })

    qualifiers.sort(key=lambda r: r['time'])
    return qualifiers


def _scan_all_qualifiers(output_dir):
    pattern = os.path.join(output_dir, 'race_data', 'race_data_*.json')
    files   = sorted(glob.glob(pattern))
    all_q   = []
    for path in files:
        date_str = os.path.basename(path).replace('race_data_', '').replace('.json', '')
        if date_str < START_DATE:
            continue
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
            all_q.extend(collect_qualifiers(data, date_str))
        except Exception as e:
            print(f'  Warning: could not read {path}: {e}')
    all_q.sort(key=lambda r: (r['date'], r['time']))
    return all_q


# ── Odds converter ────────────────────────────────────────────────────────────

def _to_decimal(price_str):
    """Convert fractional SP (e.g. '11/4') to decimal odds (e.g. 3.75). Returns float or None."""
    if not price_str or price_str == '—':
        return None
    s = re.sub(r'[FCJfcj]$', '', str(price_str).strip())
    if s.upper() in ('EVS', 'EVENS', '1/1'):
        return 2.0
    if '/' in s:
        try:
            num, den = s.split('/')
            return round(int(num) / int(den) + 1, 2)
        except Exception:
            pass
    try:
        v = float(s)
        return round(v, 2) if v > 1 else None
    except Exception:
        return None


# ── Excel writer ──────────────────────────────────────────────────────────────

def _border(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


WATCH_COLUMNS = [
    ('Date',      12),
    ('Venue',     18),
    ('Time',       8),
    ('Horse',     24),
    ('Score',      8),
    ('Gap',        8),
    ('Segment',   14),
    ('SP',        10),
    ('Decimal',    9),
    ('Result',    10),
    ('Net (£)',   10),
]
W_LAST_COL    = get_column_letter(len(WATCH_COLUMNS))
W_COL_SEGMENT = 7
W_COL_SP      = 8
W_COL_DEC     = 9
W_COL_RESULT  = 10
W_COL_NET     = 11
W_STAKE_CELL  = '$M$2'
W_CENTER      = {1, 3, 5, 6, 7, 8, 9, 10, 11}


def _write_watch_sheet(wb, watch_list, label_color):
    """Second sheet: SP < 2.0 picks — diagnostic tracking with stake calc."""
    ws = wb.create_sheet('SP<2 Watch List')

    # Title
    ws.merge_cells(f'A1:{W_LAST_COL}1')
    c = ws['A1']
    c.value     = 'SP < 2.0 Watch List  —  Diagnostic Tracking  |  No Real Bets'
    c.font      = Font(bold=True, size=13, color='f87171', name='Calibri')
    c.fill      = PatternFill('solid', fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    # Subtitle
    ws.merge_cells(f'A2:{W_LAST_COL}2')
    c = ws['A2']
    c.value     = 'Golden / Silver / Bronze qualifying picks where SP ≤ 2.0  |  Review at 30+ picks per segment'
    c.font      = Font(italic=True, size=10, color=C_SUB_FG, name='Calibri')
    c.fill      = PatternFill('solid', fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    # Stake input (M1 label, M2 value)
    lbl = ws['M1']
    lbl.value     = 'Stake (£)'
    lbl.font      = Font(bold=True, size=9, color=C_HDR_FG, name='Calibri')
    lbl.fill      = PatternFill('solid', fgColor=C_HDR_BG)
    lbl.alignment = Alignment(horizontal='center', vertical='center')

    inp = ws['M2']
    inp.value         = 10.0
    inp.font          = Font(bold=True, size=13, name='Calibri')
    inp.fill          = PatternFill('solid', fgColor=C_PRICE_BG)
    inp.alignment     = Alignment(horizontal='center', vertical='center')
    inp.border        = _border()
    inp.number_format = '£#,##0.00'
    ws.column_dimensions['M'].width = 12

    # Net balance (N1 label, N2 formula)
    bal_lbl = ws['N1']
    bal_lbl.value     = 'Net Balance'
    bal_lbl.font      = Font(bold=True, size=9, color=C_HDR_FG, name='Calibri')
    bal_lbl.fill      = PatternFill('solid', fgColor=C_HDR_BG)
    bal_lbl.alignment = Alignment(horizontal='center', vertical='center')

    bal = ws['N2']
    bal.value         = '=SUM(K5:K1048576)'
    bal.font          = Font(bold=True, size=13, name='Calibri')
    bal.alignment     = Alignment(horizontal='center', vertical='center')
    bal.border        = _border()
    bal.number_format = '[Green]+£#,##0.00;[Red]-£#,##0.00'
    ws.column_dimensions['N'].width = 14

    # Headers
    hdr_fill   = PatternFill('solid', fgColor=C_HDR_BG)
    win_fill   = PatternFill('solid', fgColor=C_WIN_BG)
    loss_fill  = PatternFill('solid', fgColor=C_LOSS_BG)
    void_fill  = PatternFill('solid', fgColor=C_VOID_BG)
    price_fill = PatternFill('solid', fgColor=C_PRICE_BG)
    alt_fill   = PatternFill('solid', fgColor=C_ALT)

    for i, (name, width) in enumerate(WATCH_COLUMNS, 1):
        c = ws.cell(row=4, column=i, value=name)
        c.font      = Font(bold=True, size=11, color=C_HDR_FG, name='Calibri')
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 22

    if not watch_list:
        ws.merge_cells(f'A5:{W_LAST_COL}5')
        c = ws['A5']
        c.value     = 'No SP<2 picks recorded yet.'
        c.font      = Font(italic=True, color=C_FOOTER, name='Calibri')
        c.alignment = Alignment(horizontal='center')
    else:
        seg_counts = {}
        for ri, w in enumerate(watch_list):
            row       = 5 + ri
            row_bg    = alt_fill if ri % 2 == 0 else None
            result    = w.get('result', '')
            price     = w.get('price', '')
            dec       = w.get('dec')
            pick_type = w.get('pick_type', 'golden')

            net_formula = (
                f'=IF(J{row}="W",I{row}*{W_STAKE_CELL}-{W_STAKE_CELL},'
                f'IF(J{row}="L",-{W_STAKE_CELL},""))'
            )
            values = [
                w['date'], w['venue'], w['time'], w['horse'],
                w['score'], w['gap'], w.get('sub_label', ''),
                price or '—',
                dec if dec is not None else '—',
                result or '—',
                net_formula,
            ]

            for ci, val in enumerate(values, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(size=11, name='Calibri')
                c.alignment = Alignment(
                    vertical='center',
                    horizontal='center' if ci in W_CENTER else 'left'
                )
                c.border = _border()

                if ci == W_COL_SEGMENT:
                    c.font = Font(size=10, bold=True, name='Calibri',
                                  color=label_color.get(pick_type, C_LABEL_GOLDEN))
                elif ci == W_COL_RESULT:
                    if result == 'W':   c.fill = win_fill
                    elif result == 'L': c.fill = loss_fill
                    elif result == 'V': c.fill = void_fill
                    else:               c.fill = price_fill
                elif ci == W_COL_NET:
                    if result == 'W':   c.fill = win_fill
                    elif result == 'L': c.fill = loss_fill
                    else:               c.fill = price_fill
                    c.number_format = '+#,##0.00;-#,##0.00'
                elif ci in (W_COL_SP, W_COL_DEC):
                    c.fill = price_fill
                elif row_bg:
                    c.fill = row_bg

            ws.row_dimensions[row].height = 20
            seg_counts[pick_type] = seg_counts.get(pick_type, 0) + 1

        # Footer summary per segment
        footer_row = 5 + len(watch_list) + 1
        ws.merge_cells(f'A{footer_row}:{W_LAST_COL}{footer_row}')
        c = ws[f'A{footer_row}']
        parts = [f'{k.capitalize()}: {v}' for k, v in sorted(seg_counts.items())]
        c.value     = 'SP<2 totals:  ' + '  |  '.join(parts) + f'  —  Updated {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
        c.font      = Font(italic=True, size=9, color=C_FOOTER, name='Calibri')
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[footer_row].height = 16

    ws.freeze_panes = 'A5'


def _write_combined_excel(all_qualifiers, history_lookup, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Qualifying Picks'

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{LAST_COL}1')
    c = ws['A1']
    c.value = 'Prixm Qualifying Picks  —  May 2026 onwards'
    c.font      = Font(bold=True, size=14, color=C_TITLE_FG, name='Calibri')
    c.fill      = PatternFill('solid', fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Subtitle ──────────────────────────────────────────────────────────────
    ws.merge_cells(f'A2:{LAST_COL}2')
    c = ws['A2']
    c.value = 'Gap >= 17  |  Non-Handicap  |  England only  |  Result + Price auto-filled from results history'
    c.font      = Font(italic=True, size=10, color=C_SUB_FG, name='Calibri')
    c.fill      = PatternFill('solid', fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6   # spacer

    # ── Stake input (N1 label, N2 value — user edits N2) ─────────────────────
    lbl = ws['O1']
    lbl.value     = 'Stake (£)'
    lbl.font      = Font(bold=True, size=9, color=C_HDR_FG, name='Calibri')
    lbl.fill      = PatternFill('solid', fgColor=C_HDR_BG)
    lbl.alignment = Alignment(horizontal='center', vertical='center')

    inp = ws['O2']
    inp.value          = 10.0
    inp.font           = Font(bold=True, size=13, name='Calibri')
    inp.fill           = PatternFill('solid', fgColor=C_PRICE_BG)
    inp.alignment      = Alignment(horizontal='center', vertical='center')
    inp.border         = _border()
    inp.number_format  = '£#,##0.00'
    ws.column_dimensions['O'].width = 12

    bal_lbl = ws['P1']
    bal_lbl.value     = 'Net Balance'
    bal_lbl.font      = Font(bold=True, size=9, color=C_HDR_FG, name='Calibri')
    bal_lbl.fill      = PatternFill('solid', fgColor=C_HDR_BG)
    bal_lbl.alignment = Alignment(horizontal='center', vertical='center')

    bal = ws['P2']
    bal.value         = '=SUM(M5:M1048576)'
    bal.font          = Font(bold=True, size=13, name='Calibri')
    bal.alignment     = Alignment(horizontal='center', vertical='center')
    bal.border        = _border()
    bal.number_format = '[Green]+£#,##0.00;[Red]-£#,##0.00'
    ws.column_dimensions['P'].width = 14

    # ── Headers ───────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill('solid', fgColor=C_HDR_BG)
    price_fill = PatternFill('solid', fgColor=C_PRICE_BG)
    win_fill   = PatternFill('solid', fgColor=C_WIN_BG)
    loss_fill  = PatternFill('solid', fgColor=C_LOSS_BG)
    void_fill  = PatternFill('solid', fgColor=C_VOID_BG)
    alt_fill   = PatternFill('solid', fgColor=C_ALT)

    label_color = {
        'golden': C_LABEL_GOLDEN,
        'silver': C_LABEL_SILVER,
        'bronze': C_LABEL_BRONZE,
    }

    for i, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=4, column=i, value=name)
        c.font      = Font(bold=True, size=11, color=C_HDR_FG, name='Calibri')
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    center_cols = {COL_DATE, COL_TIME, COL_SCORE, COL_GAP, COL_SUB_LABEL, COL_PRICE, COL_DECIMAL, COL_RESULT, COL_NET}

    rendered        = 0
    golden_results  = []
    silver_results  = []
    bronze_results  = []
    watch_list      = []   # SP < 2.0 entries — written to separate sheet

    if not all_qualifiers:
        ws.merge_cells(f'A5:{LAST_COL}5')
        c = ws['A5']
        c.value     = 'No qualifying horses found from May 2026 onwards.'
        c.font      = Font(italic=True, color=C_FOOTER, name='Calibri')
        c.alignment = Alignment(horizontal='center')
    else:
        for q in all_qualifiers:
            result, price = _resolve_result(q, history_lookup)
            dec = _to_decimal(price)

            # Price known and below 2.00 — divert to watch list, skip main sheet
            if dec is not None and dec < 2.0:
                watch_list.append({**q, 'result': result, 'price': price, 'dec': dec})
                continue

            pick_type = q.get('pick_type', 'golden')
            row       = 5 + rendered
            row_bg    = alt_fill if rendered % 2 == 0 else None
            rendered += 1

            net_formula = (
                f'=IF(L{row}="W",K{row}*{STAKE_CELL}-{STAKE_CELL},'
                f'IF(L{row}="L",-{STAKE_CELL},""))'
            )
            values = [
                q['date'], q['venue'], q['time'], q['horse'],
                q['score'], q['gap'], q.get('sub_label', ''),
                q['jockey'], q['trainer'],
                price or '—',
                dec if dec is not None else '—',
                result or '—',
                net_formula,
            ]

            for ci, val in enumerate(values, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(size=11, name='Calibri')
                c.alignment = Alignment(
                    vertical='center',
                    horizontal='center' if ci in center_cols else 'left'
                )
                c.border = _border()

                if ci == COL_SUB_LABEL:
                    c.font = Font(size=10, bold=True, name='Calibri',
                                  color=label_color.get(pick_type, C_LABEL_GOLDEN))
                elif ci == COL_RESULT:
                    if result == 'W':
                        c.fill = win_fill
                    elif result == 'L':
                        c.fill = loss_fill
                    elif result == 'V':
                        c.fill = void_fill
                    else:
                        c.fill = price_fill
                elif ci == COL_NET:
                    if result == 'W':
                        c.fill = win_fill
                    elif result == 'L':
                        c.fill = loss_fill
                    else:
                        c.fill = price_fill
                    c.number_format = '+#,##0.00;-#,##0.00'
                elif ci in (COL_PRICE, COL_DECIMAL):
                    c.fill = price_fill
                elif row_bg:
                    c.fill = row_bg

            ws.row_dimensions[row].height = 20
            if pick_type == 'silver':
                silver_results.append(result)
            elif pick_type == 'bronze':
                bronze_results.append(result)
            else:
                golden_results.append(result)

    # ── Result dropdown (W/L/V) ───────────────────────────────────────────────
    n_data = max(rendered, 1)
    res_letter = get_column_letter(COL_RESULT)
    dv = DataValidation(
        type='list', formula1='"W,L,V"',
        allow_blank=True, showDropDown=False,
    )
    dv.sqref = f'{res_letter}5:{res_letter}{5 + n_data - 1}'
    ws.add_data_validation(dv)

    # ── Footer ────────────────────────────────────────────────────────────────
    def _sr(results):
        w = results.count('W'); l = results.count('L'); d = w + l
        return f'W:{w}  L:{l}  SR:{round(w/d*100)}%' if d else 'no results yet'

    n_excl    = len(watch_list)
    excl_note = f'  |  {n_excl} on SP<2 Watch List' if n_excl else ''
    target    = 30

    def _paper_note(results, target):
        n = len(results)
        return f'PAPER {n}/{target}' if n < target else f'PAPER {n} ✓'

    footer_row = 5 + max(rendered, 1) + 1
    ws.merge_cells(f'A{footer_row}:{LAST_COL}{footer_row}')
    c = ws[f'A{footer_row}']
    c.value = (
        f'Golden ({len(golden_results)} rows): {_sr(golden_results)}{excl_note}'
        f'    |    '
        f'Silver ({_paper_note(silver_results, target)}): {_sr(silver_results)}'
        f'    |    '
        f'Bronze ({_paper_note(bronze_results, target)}): {_sr(bronze_results)}'
        f'    —    Updated {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )
    c.font      = Font(italic=True, size=9, color=C_FOOTER, name='Calibri')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[footer_row].height = 16

    ws.freeze_panes = 'A5'

    _write_watch_sheet(wb, watch_list, label_color)

    wb.save(out_path)


# ── Public API ────────────────────────────────────────────────────────────────

def rebuild_qualifying_excel(output_dir):
    """
    Full rebuild: scan all race_data/*.json >= START_DATE,
    cross-reference results_history.json, write qualifying_picks.xlsx.
    Returns (out_path, shown_count, resolved_count) — shown excludes SP < 2.00.
    """
    os.makedirs(os.path.join(output_dir, 'output'), exist_ok=True)
    out_path       = os.path.join(output_dir, 'output', 'qualifying_picks.xlsx')
    all_qualifiers = _scan_all_qualifiers(output_dir)
    history_lookup = _build_history_lookup(output_dir)
    _write_combined_excel(all_qualifiers, history_lookup, out_path)
    # Count only rows that pass the price filter (mirrors _write_combined_excel logic)
    shown    = 0
    resolved = 0
    for q in all_qualifiers:
        result, price = _resolve_result(q, history_lookup)
        dec = _to_decimal(price)
        if dec is not None and dec < 2.0:
            continue
        shown += 1
        if result:
            resolved += 1
    return out_path, shown, resolved


def generate_qualifying_excel(new_data, date_str, output_dir):
    """
    Called by racecard fetchers after saving JSON.
    Rebuilds combined file and returns (out_path, new_date_qualifier_count).
    """
    out_path, total, resolved = rebuild_qualifying_excel(output_dir)
    new_count = sum(
        1 for q in _scan_all_qualifiers(output_dir) if q['date'] == date_str
    )
    return out_path, new_count


# ── Standalone entry point ────────────────────────────────────────────────────

def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='Generate Prixm qualifying picks Excel')
    parser.add_argument('--scan', action='store_true',
                        help='Rebuild combined file (default behaviour)')
    args = parser.parse_args()

    print('=== Prixm Qualifying Exporter ===\n')

    all_qualifiers = _scan_all_qualifiers(OUTPUT_DIR)
    history_lookup = _build_history_lookup(OUTPUT_DIR)

    def _console_sr(qs, lookup):
        results = [_resolve_result(q, lookup)[0] for q in qs]
        w = results.count('W'); l = results.count('L'); d = w + l
        p = len(qs) - d
        sr = f'{round(w/d*100)}%' if d else '—'
        return w, l, d, p, sr

    if not all_qualifiers:
        print(f'No qualifying horses found from {START_DATE} onwards.')
    else:
        golden_qs = [q for q in all_qualifiers if q.get('pick_type') == 'golden']
        silver_qs = [q for q in all_qualifiers if q.get('pick_type') == 'silver']
        bronze_qs = [q for q in all_qualifiers if q.get('pick_type') == 'bronze']
        dates     = sorted({q['date'] for q in all_qualifiers})

        gw, gl, gd, gp, gsr = _console_sr(golden_qs, history_lookup)
        sw, sl, sd, sp, ssr = _console_sr(silver_qs, history_lookup)
        bw, bl, bd, bp, bsr = _console_sr(bronze_qs, history_lookup)

        print(f'Dates:          {len(dates)}  ({dates[0]} to {dates[-1]})')
        print(f'Golden:         {len(golden_qs)} rows  |  Results: {gd}  |  Pending: {gp}  |  W:{gw}  L:{gl}  SR:{gsr}')
        print(f'Silver (paper): {len(silver_qs)} rows  |  Results: {sd}/30  |  Pending: {sp}  |  W:{sw}  L:{sl}  SR:{ssr}')
        print(f'Bronze (paper): {len(bronze_qs)} rows  |  Results: {bd}/30  |  Pending: {bp}  |  W:{bw}  L:{bl}  SR:{bsr}\n')

        pfx_map = {'silver': '(S) ', 'bronze': '(B) '}
        for d in dates:
            qs    = [q for q in all_qualifiers if q['date'] == d]
            names = []
            for q in qs:
                r, _ = _resolve_result(q, history_lookup)
                tag  = f'[{r}]' if r else '[ ]'
                pfx  = pfx_map.get(q.get('pick_type', ''), '')
                names.append(f'{pfx}{q["horse"]} {tag}')
            print(f'  {d}: {", ".join(names)}')

    _write_combined_excel(all_qualifiers, history_lookup, OUTPUT_FILE)
    print(f'\nOutput: {OUTPUT_FILE}')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'\n!!! CRASH: {e}')
        import traceback
        traceback.print_exc()
    print()
    input('Press Enter to close...')
